#!/usr/bin/env python
# usage: ./convert_to_csv.py <input_file> <output_file>

import csv
import re
import sys

from openpyxl import load_workbook


if len(sys.argv) != 3:
    print("usage: ./convert_to_csv.py <input_file> <output_file>")
    sys.exit(1)

input_file = sys.argv[1]
output_file = sys.argv[2]

# 엑셀 불러오기
wb = load_workbook(input_file, data_only=True)
ws = wb.active  # 첫 번째 시트 사용

# 제외할 패턴 (정규식)
exclude_pattern = re.compile(r"(이표기|오표기|공개\s*여부|출전)")

# 헤더 행 추출
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
# 남길 열 인덱스 결정
keep_indices = [i for i, h in enumerate(headers) if h and not exclude_pattern.search(str(h))]

# CSV로 저장
with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    # 헤더 쓰기
    writer.writerow([headers[i] for i in keep_indices])
    # 데이터 쓰기
    for row in ws.iter_rows(min_row=2, values_only=True):
        writer.writerow([row[i] for i in keep_indices])
